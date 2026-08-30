"""Safe extraction of visible content from macro-enabled Excel workbooks.

An XLSM file is an untrusted ZIP archive.  This module validates the archive before
passing it to :mod:`openpyxl`, and deliberately never loads VBA or resolves external
links.  The extractor returns the shared ``ExtractionRecord`` contract used by the
rest of the ingestion pipeline.
"""

from __future__ import annotations

import hashlib
import re
import zipfile
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any

from defusedxml import ElementTree as DefusedET
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from scripts.ingestion.models import ExtractionMethod, ExtractionRecord, WarningCode

# These limits are intentionally conservative for downloaded government forms.  They
# bound metadata inspection as well as the eventual openpyxl parse.
MAX_ARCHIVE_UNCOMPRESSED = 100 * 1024 * 1024
MAX_MEMBER_UNCOMPRESSED = 25 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 10_000
MAX_VALIDATION_CELLS = 20_000
_ZIP_SIGNATURES = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
_REQUIRED_MEMBERS = frozenset({"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"})
_HTML_MARKERS = (b"<html", b"<!doctype html", b"<head", b"<body")


class XLSMExtractionError(ValueError):
    """Raised when a file is not a safe, structurally valid XLSM workbook."""


@dataclass(frozen=True)
class ArchiveInfo:
    """Validated archive metadata, useful to callers performing a preflight."""

    member_count: int
    uncompressed_size: int
    members: tuple[str, ...]


@dataclass
class _ValidationCapture:
    text: str
    partial: bool = False


def _looks_like_html(data: bytes) -> bool:
    """Detect common HTML responses before any workbook parser is invoked."""

    sample = data[:8192].lstrip().lower()
    return any(sample.startswith(marker) or marker in sample[:1024] for marker in _HTML_MARKERS)


def _validate_member_name(name: str) -> None:
    """Reject absolute and traversal paths, including Windows-style variants."""

    if not name or "\x00" in name or "\\" in name:
        raise XLSMExtractionError("XLSM archive contains an unsafe member name")
    path = PurePosixPath(name)
    if path.is_absolute() or ".." in path.parts:
        raise XLSMExtractionError("XLSM archive contains a path-traversal member")


def validate_xlsm_archive(path: str | Path) -> ArchiveInfo:
    """Validate ZIP signature, required workbook members, names, and size limits.

    No archive member is extracted to disk.  XML parts are parsed with defusedxml to
    reject entity-expansion and related XML attacks before openpyxl gets the file.
    """

    workbook_path = Path(path)
    try:
        with workbook_path.open("rb") as stream:
            signature = stream.read(8)
            if not signature.startswith(_ZIP_SIGNATURES):
                if _looks_like_html(signature + stream.read(8192)):
                    raise XLSMExtractionError("XLSM download is HTML, not an Excel workbook")
                raise XLSMExtractionError("XLSM file does not have a ZIP signature")
            stream.seek(0)
            if _looks_like_html(stream.read(8192)):
                raise XLSMExtractionError("XLSM download is HTML, not an Excel workbook")
    except OSError as exc:
        raise XLSMExtractionError(f"cannot read XLSM file: {workbook_path}") from exc

    try:
        with zipfile.ZipFile(workbook_path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_MEMBERS:
                raise XLSMExtractionError("XLSM archive contains too many members")
            total_size = 0
            names: list[str] = []
            for info in infos:
                _validate_member_name(info.filename)
                if info.file_size > MAX_MEMBER_UNCOMPRESSED:
                    raise XLSMExtractionError(f"XLSM archive member is too large: {info.filename}")
                total_size += info.file_size
                if total_size > MAX_ARCHIVE_UNCOMPRESSED:
                    raise XLSMExtractionError("XLSM archive has an unreasonable uncompressed size")
                names.append(info.filename)

            present = set(names)
            if len(present) != len(names):
                raise XLSMExtractionError("XLSM archive contains duplicate member names")
            missing = _REQUIRED_MEMBERS - present
            if missing:
                raise XLSMExtractionError(
                    f"XLSM archive is missing required member(s): {', '.join(sorted(missing))}"
                )
            worksheet_members = [name for name in names if name.startswith("xl/worksheets/")]
            if not worksheet_members:
                raise XLSMExtractionError("XLSM archive contains no worksheet XML parts")

            # Parsing only the metadata parts verifies that entity-bearing XML cannot
            # reach openpyxl.  defusedxml raises a DefusedXmlException on unsafe XML.
            for member in ("[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"):
                try:
                    DefusedET.fromstring(archive.read(member))
                except Exception as exc:  # defusedxml has version-specific exception classes
                    raise XLSMExtractionError(f"invalid or unsafe XML member: {member}") from exc
    except (OSError, zipfile.BadZipFile) as exc:
        raise XLSMExtractionError("XLSM file is not a valid ZIP archive") from exc

    return ArchiveInfo(len(infos), total_size, tuple(names))


def _normalise_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return (
        str(value)
        # Excel serializes line breaks and a few XML controls as escaped tokens in
        # some downloaded forms; turn them back into the visible text characters.
        .replace("_x000a_", "\n")
        .replace("_x000d_", "\n")
        .replace("_x0009_", "\t")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .strip()
    )


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _validation_formula_values(formula: str | None) -> tuple[str, ...] | None:
    """Return literal list-validation values; None means not safely enumerable."""

    if not formula:
        return None
    value = formula.strip()
    if len(value) >= 2 and value[0] == value[-1] == '"':
        return tuple(item.strip() for item in value[1:-1].split(",") if item.strip())
    return None


def _validation_text(validation: DataValidation) -> _ValidationCapture:
    """Capture user-facing validation text without evaluating workbook formulas."""

    fragments: list[str] = []
    if validation.promptTitle:
        fragments.append(f"Prompt title: {_normalise_text(validation.promptTitle)}")
    if validation.prompt:
        fragments.append(f"Prompt: {_normalise_text(validation.prompt)}")
    if validation.errorTitle:
        fragments.append(f"Error title: {_normalise_text(validation.errorTitle)}")
    if validation.error:
        fragments.append(f"Error: {_normalise_text(validation.error)}")

    partial = False
    values = _validation_formula_values(validation.formula1)
    if values:
        fragments.append("Allowed values: " + ", ".join(values))
    elif validation.type == "list" and validation.formula1:
        # A range or named formula is intentionally not evaluated.  It may point to
        # another workbook, and resolving it would violate the no-external-links rule.
        partial = True
    elif validation.type not in {None, "list"}:
        partial = True
    if validation.type == "list" and not validation.formula1:
        partial = True

    return _ValidationCapture("; ".join(fragments), partial)


def _iter_validation_cells(
    worksheet: Any, validation: DataValidation
) -> Iterable[tuple[int, int, bool]]:
    """Yield bounded validation cells as (row, column, was_truncated)."""

    remaining = MAX_VALIDATION_CELLS
    for cell_range in validation.sqref.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
        except ValueError:
            yield 0, 0, True
            continue
        # A whole-column/whole-row validation can otherwise expand to 16k x 1m.
        max_row = min(max_row, max(worksheet.max_row, min_row))
        max_col = min(max_col, max(worksheet.max_column, min_col))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if remaining <= 0:
                    yield 0, 0, True
                    return
                yield row, col, False
                remaining -= 1


def _cell_display(cell: Any, validation: str = "") -> str:
    value = _normalise_text(cell.value)
    if not value and validation:
        value = "[blank input]"
    elif not value and cell.comment and cell.comment.text:
        value = "[blank cell]"
    if not value:
        return ""
    pieces = [value]
    if cell.comment and cell.comment.text:
        pieces.append(f"Comment: {_normalise_text(cell.comment.text)}")
    if validation:
        pieces.append(validation)
    number_format = _normalise_text(cell.number_format)
    if number_format and number_format != "General":
        # Include formats only when they communicate an expected input shape.
        lowered = number_format.lower()
        if any(token in lowered for token in ("yy", "dd", "%", "$", "£", "€", "#,##0")):
            pieces.append(f"Expected format: {number_format}")
    return " | ".join(pieces)


def _slug_fragment(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "sheet"


def _source_id_for(path: Path, source_id: str | None) -> str:
    if source_id:
        return source_id
    candidate = _slug_fragment(path.stem)
    return candidate[:100] if len(candidate) >= 3 else "xlsm-source"


def _row_regions(rows: list[int]) -> list[tuple[int, int]]:
    if not rows:
        return []
    regions: list[tuple[int, int]] = []
    start = previous = rows[0]
    for row in rows[1:]:
        if row != previous + 1:
            regions.append((start, previous))
            start = row
        previous = row
    regions.append((start, previous))
    return regions


def _heading_context(worksheet: Any, first_row: int) -> str:
    candidates: list[str] = []
    for row in range(max(1, first_row - 5), first_row):
        values = [
            _normalise_text(cell.value)
            for (cell_row, _), cell in worksheet._cells.items()  # noqa: SLF001
            if cell_row == row
            if not isinstance(cell, MergedCell) and _normalise_text(cell.value)
        ]
        if len(values) == 1 and len(values[0]) <= 160:
            candidates.append(values[0])
    return candidates[-1] if candidates else ""


def _safe_defined_name_destinations(workbook: Any) -> dict[str, list[tuple[str, str]]]:
    """Return local worksheet destinations without resolving external formulas.

    ``openpyxl`` exposes defined names through a convenience ``destinations``
    iterator, but a name may also contain an external workbook reference or a
    formula (``OFFSET``, ``INDIRECT``, and so on).  Those values are metadata only
    and must never be evaluated.  We retain only finite, local A1 ranges that can
    be read from the already-loaded worksheet.
    """

    result: dict[str, list[tuple[str, str]]] = {}
    sheet_names = set(workbook.sheetnames)
    defined_names = getattr(workbook, "defined_names", {})
    values = defined_names.values() if hasattr(defined_names, "values") else ()
    for defined_name in values:
        name = _normalise_text(getattr(defined_name, "name", None))
        if not name or name.lower().startswith("_xlnm."):
            continue
        attr_text = _normalise_text(getattr(defined_name, "attr_text", None))
        if not attr_text or any(token in attr_text for token in ("[", "]", "http:", "https:")):
            continue
        try:
            destinations = list(defined_name.destinations)
        except (TypeError, ValueError, AttributeError):
            continue
        for sheet_name, coordinate in destinations:
            sheet_name = str(sheet_name)
            coordinate = str(coordinate)
            if (
                sheet_name not in sheet_names
                or ":" in sheet_name
                or "!" in coordinate
                or "[" in coordinate
                or "]" in coordinate
            ):
                continue
            try:
                min_col, min_row, max_col, max_row = range_boundaries(coordinate)
            except ValueError:
                continue
            if min_col < 1 or min_row < 1 or max_col < min_col or max_row < min_row:
                continue
            if (max_col - min_col + 1) * (max_row - min_row + 1) > MAX_VALIDATION_CELLS:
                # Whole-column named ranges are metadata for Excel, not useful
                # corpus content; never expand them into millions of cells.
                continue
            canonical = (
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            )
            destination = (name, canonical)
            if destination not in result.setdefault(sheet_name, []):
                result[sheet_name].append(destination)
    return result


def _defined_range_record(
    worksheet: Any,
    source_id: str,
    start_ordinal: int,
    name: str,
    cell_range: str,
    relevant: Mapping[tuple[int, int], Any],
    validation_by_cell: Mapping[tuple[int, int], list[str]],
    validation_partial: set[tuple[int, int]],
) -> ExtractionRecord:
    """Build one record for a safe, local defined range."""

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    lines: list[str] = []
    for row in range(min_row, max_row + 1):
        row_parts: list[str] = []
        for col in range(min_col, max_col + 1):
            cell = relevant.get((row, col))
            if cell is None:
                continue
            validation = "; ".join(validation_by_cell.get((row, col), []))
            if not validation and (row, col) in validation_partial:
                validation = "Validation present (details not expanded)"
            display = _cell_display(cell, validation)
            if display:
                row_parts.append(f"{cell.coordinate}: {display}")
        if row_parts:
            lines.append(" | ".join(row_parts))
    content_parts = [
        f"Sheet: {worksheet.title}",
        f"Named range: {name}",
        f"Range: {cell_range}",
    ]
    if lines:
        content_parts.extend(lines)
    else:
        content_parts.append("[named range contains no visible cell values]")
    content = "\n".join(content_parts)
    digest = hashlib.sha256(content.encode("utf-8")).hexdigest()
    return ExtractionRecord(
        record_id=f"{source_id}-{_slug_fragment(worksheet.title)}-named-{_slug_fragment(name)}-{digest[:12]}",
        source_id=source_id,
        content_kind="worksheet",
        ordinal=start_ordinal,
        title_path=[worksheet.title, f"Named range: {name}"],
        content=content,
        section=f"Named range: {name}",
        sheet=worksheet.title,
        cell_range=cell_range,
        extraction_method=ExtractionMethod.XLSM_XML,
        warnings=[],
        content_hash=digest,
    )


def _extract_sheet(
    worksheet: Any,
    source_id: str,
    start_ordinal: int,
    inherited_warnings: list[WarningCode],
    defined_ranges: Iterable[tuple[str, str]] = (),
) -> tuple[list[ExtractionRecord], int]:
    validation_by_cell: dict[tuple[int, int], list[str]] = {}
    validation_partial: set[tuple[int, int]] = set()
    for validation in worksheet.data_validations.dataValidation:
        capture = _validation_text(validation)
        for row, col, truncated in _iter_validation_cells(worksheet, validation):
            if row == 0:
                # Malformed or over-sized ranges are reported on an existing output
                # record below; never try to instantiate arbitrary cells.
                validation_partial.add((-1, -1))
                continue
            validation_by_cell.setdefault((row, col), [])
            if capture.text:
                validation_by_cell[(row, col)].append(capture.text)
            if capture.partial or truncated:
                validation_partial.add((row, col))

    relevant: dict[tuple[int, int], Any] = {}
    for (row, col), cell in worksheet._cells.items():  # noqa: SLF001 - avoids formatting-only cells
        if isinstance(cell, MergedCell):
            continue
        if cell.value is not None or cell.comment is not None:
            relevant[(row, col)] = cell
    for row, col in validation_by_cell:
        cell = worksheet.cell(row=row, column=col)
        relevant[(row, col)] = cell

    rows = sorted({row for row, _ in relevant})
    records: list[ExtractionRecord] = []
    warnings = list(dict.fromkeys(inherited_warnings))
    for first_row, last_row in _row_regions(rows):
        region_cells = {
            (row, col): cell
            for (row, col), cell in relevant.items()
            if first_row <= row <= last_row
        }
        if not region_cells:
            continue
        min_col = min(col for _, col in region_cells)
        max_col = max(col for _, col in region_cells)
        cell_range = (
            f"{get_column_letter(min_col)}{first_row}:{get_column_letter(max_col)}{last_row}"
        )
        lines: list[str] = []
        formula_seen = False
        region_warnings = list(warnings)
        for row in range(first_row, last_row + 1):
            row_parts: list[str] = []
            for col in range(min_col, max_col + 1):
                cell = region_cells.get((row, col))
                if cell is None:
                    continue
                validation = "; ".join(validation_by_cell.get((row, col), []))
                if not validation and (row, col) in validation_partial:
                    validation = "Validation present (details not expanded)"
                display = _cell_display(cell, validation)
                if not display:
                    continue
                row_parts.append(f"{cell.coordinate}: {display}")
                formula_seen = formula_seen or _is_formula(cell.value)
                if (row, col) in validation_partial or (-1, -1) in validation_partial:
                    if WarningCode.DATA_VALIDATION_PARTIAL not in region_warnings:
                        region_warnings.append(WarningCode.DATA_VALIDATION_PARTIAL)
            if row_parts:
                lines.append(" | ".join(row_parts))
        if not lines:
            continue
        if formula_seen and WarningCode.FORMULA_WITHOUT_CACHED_VALUE not in region_warnings:
            region_warnings.append(WarningCode.FORMULA_WITHOUT_CACHED_VALUE)
        heading = _heading_context(worksheet, first_row)
        content_parts = [f"Sheet: {worksheet.title}"]
        if heading:
            content_parts.append(f"Heading: {heading}")
        content_parts.extend(lines)
        content = "\n".join(content_parts)
        digest = hashlib.sha256(content.encode("utf-8")).hexdigest()
        title_path = [worksheet.title]
        if heading:
            title_path.append(heading)
        records.append(
            ExtractionRecord(
                record_id=f"{source_id}-{_slug_fragment(worksheet.title)}-{digest[:16]}",
                source_id=source_id,
                content_kind="worksheet",
                ordinal=start_ordinal + len(records),
                title_path=title_path,
                content=content,
                section=heading or worksheet.title,
                sheet=worksheet.title,
                cell_range=cell_range,
                extraction_method=ExtractionMethod.XLSM_XML,
                warnings=region_warnings,
                content_hash=digest,
            )
        )
    # Defined ranges are emitted independently from logical row regions.  This
    # preserves the exact named-range locator even where it overlaps a larger
    # table, while keeping the ordinary sheet extraction deterministic.
    for name, cell_range in defined_ranges:
        records.append(
            _defined_range_record(
                worksheet,
                source_id,
                start_ordinal + len(records),
                name,
                cell_range,
                relevant,
                validation_by_cell,
                validation_partial,
            )
        )
    return records, start_ordinal + len(records)


def extract_xlsm(path: str | Path, source_id: str | None = None) -> list[ExtractionRecord]:
    """Extract visible, non-empty worksheet regions from a local XLSM file.

    ``keep_vba=False`` and ``keep_links=False`` ensure macro projects and external
    workbook relationships are neither loaded nor resolved.  Formula cells remain
    formula text because ``data_only=False`` is required by the extraction contract.
    """

    workbook_path = Path(path)
    validate_xlsm_archive(workbook_path)
    resolved_source_id = _source_id_for(workbook_path, source_id)
    try:
        workbook = load_workbook(
            workbook_path,
            read_only=False,
            data_only=False,
            keep_vba=False,
            keep_links=False,
        )
    except Exception as exc:
        raise XLSMExtractionError("openpyxl could not parse the validated XLSM workbook") from exc

    hidden_count = sum(worksheet.sheet_state != "visible" for worksheet in workbook.worksheets)
    inherited = [WarningCode.HIDDEN_SHEET_SKIPPED] if hidden_count else []
    defined_ranges = _safe_defined_name_destinations(workbook)
    records: list[ExtractionRecord] = []
    ordinal = 1
    try:
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            sheet_records, ordinal = _extract_sheet(
                worksheet,
                resolved_source_id,
                ordinal,
                inherited,
                defined_ranges.get(worksheet.title, ()),
            )
            records.extend(sheet_records)
    finally:
        workbook.close()
    return records


# Friendly aliases for orchestration code and callers that use noun-oriented names.
extract_workbook = extract_xlsm
extract = extract_xlsm

__all__ = [
    "ArchiveInfo",
    "MAX_ARCHIVE_MEMBERS",
    "MAX_ARCHIVE_UNCOMPRESSED",
    "MAX_MEMBER_UNCOMPRESSED",
    "XLSMExtractionError",
    "extract",
    "extract_workbook",
    "extract_xlsm",
    "validate_xlsm_archive",
]
